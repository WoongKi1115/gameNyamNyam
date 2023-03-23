import requests
from fastapi import FastAPI, Depends
from fastapi.responses import JSONResponse
from mongoengine import connect
from openpyxl import load_workbook
from starlette.requests import Request
from steamsignin import SteamSignIn
import models
import json

app = FastAPI()
# db 연동
connect(db="nyamnyam", host="localhost", port=27017)
api_url = "http://127.0.0.1:8000"


@app.get("/games/{user_id}")
async def get_game_count(user_id: str):
    url = f"https://api.steampowered.com/IPlayerService/GetOwnedGames/v0001/?key=FF1F5EF115E50CEBE0A7C11123959310&steamid={user_id}&format=json"
    response = requests.get(url)
    json_response = response.json()
    game_count = json_response["response"]["game_count"]
    return game_count


@app.get('/login')
async def main(steam_signin: SteamSignIn = Depends(SteamSignIn)):
    url = steam_signin.ConstructURL(api_url+'/processlogin')
    return steam_signin.RedirectUser(url)


@app.get('/processlogin')
async def pr(request: Request, steam_signin: SteamSignIn = Depends(SteamSignIn)):
    return steam_signin.ValidateResults(request.query_params)
# 규칙에 따른 추천 게임 리스트 (인기 게임 몇 개, 사용자 리스트 기반 몇 개 ...)


@app.get("/games/all")
def get_all_game():
    return {"Hello": "World"}

# 5개 이상 플레이한 사용자가 전에 했던 게임 기반으로 (태그) 분석


@app.get("/games/result/{userid}")
def get_game_by_tag():
    return {"Hello": "World"}

# 5개 미만 플레이한 사용자의 장바구니 기반으로 (태그) 분석


@app.get("/games/result")
def get_game_by_table(table_list: str):
    return {"Hello": "World"}

# 뽑은 게임 리스트에 대한 매치율 저장


@app.get("/games/rate")
def get_rate():
    return {"Hello": "World"}

# 뽑은 리스트랑 비슷한 거 추천


@app.get("/games/similar")
def get_similar_game():

    return {"Hello": "World"}

# 저장된 game 데이터 전체 가져오기
@app.get("/get_game")
def get_all_game():
    # 객체 가져오기
    games = models.Game.objects()
    # return list(games)

    # json으로 형변환 후 return
    games_json = games.to_json()
    return JSONResponse(json.loads(games_json))

# 엑셀파일 불러와서 데이터 넣기
@app.get("/insert_excel")
def excel_to_db():
    print("insert data...")

    # 엑셀파일 열기
    excel_filename = "./game_data_50000.xlsx"
    wb = load_workbook(excel_filename)
    ws = wb.active

    for x in range(2, ws.max_row):
        # 엑셀 파일에서 하나씩 데이터 뽑아오기
        appid = ws.cell(row=x, column=2).value
        name = str(ws.cell(row=x,column=3).value)
        short_description = str(ws.cell(row=x,column=4).value)
        price = ws.cell(row=x, column=5).value
        categories = ws.cell(row=x, column=6).value
        genres = ws.cell(row=x, column=7).value
        recommendations = ws.cell(row=x,column=8).value
        release_date = ws.cell(row=x, column=9).value
        developers = str(ws.cell(row=x, column=10).value)
        metacritic = ws.cell(row=x, column=11).value
        image = str(ws.cell(row=x, column=12).value)
        about_the_game = str(ws.cell(row=x, column=13).value)
        screenshots = str(ws.cell(row=x, column=14).value)

        # str split -> 리스트로 변환 
        categories_list = categories.split(",")
        genres_list = genres.split(",")

        g = models.Game(appid=appid, name=name, short_description=short_description,price=price,recommendations=recommendations,release_date=release_date,developers=developers,metacritic=metacritic,image=image,about_the_game=about_the_game,screenshots=screenshots)
        g.categories = categories_list
        g.genres = genres_list

        g.save() # 저장하기
        


